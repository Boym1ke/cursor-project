"""
Author: Cassius Mai
Description: This script processes TOPS_Costs preprocessing.xlsx and TOPS_Costs url.xlsx files,
             and updates readurl.txt with the results.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os
from collections import defaultdict

def process_tops_costs_file(preprocessing_file, url_file, readurl_file):
    # 检查文件是否存在
    if not os.path.exists(preprocessing_file) or not os.path.exists(url_file):
        print(f"错误: 文件不存在。")
        return

    # 处理 TOPS_Costs preprocessing.xlsx
    wb_preprocessing = openpyxl.load_workbook(preprocessing_file)
    sheet_preprocessing = wb_preprocessing.active

    # 找到最后一行
    max_row = sheet_preprocessing.max_row
    for row in range(max_row, 2, -1):
        if sheet_preprocessing[f'A{row}'].value is not None:
            max_row = row
            break

    # 存储需要删除的行和Detention相关的A列值
    rows_to_delete = set()
    detention_a_values = set()
    a_column_values = defaultdict(list)

    # 第一次遍历：标记需要删除的行和收集Detention相关的A列值
    for row in range(max_row, 2, -1):  # 从第3行开始,因为A3是起始点
        a_value = sheet_preprocessing[f'A{row}'].value
        b_value = sheet_preprocessing[f'B{row}'].value
        d_value = sheet_preprocessing[f'D{row}'].value
        p_value = str(sheet_preprocessing[f'P{row}'].value) if sheet_preprocessing[f'P{row}'].value else ''

        if b_value is not None:
            rows_to_delete.add(row)
        elif d_value is not None and 'Detention' not in p_value:
            rows_to_delete.add(row)
        
        if 'Detention' in p_value:
            detention_a_values.add(a_value)
            rows_to_delete.add(row)

        # 收集A列的值和对应的行号
        a_column_values[a_value].append(row)

    # 第二次遍历：标记与Detention相关的A列值相同的行
    for row in range(max_row, 2, -1):
        a_value = sheet_preprocessing[f'A{row}'].value
        if a_value in detention_a_values:
            rows_to_delete.add(row)

    # 处理A列重复数据
    for a_value, rows in a_column_values.items():
        if len(rows) > 1:  # 如果有重复
            rows_to_delete.update(rows[1:])  # 保留第一个出现的行，删除其余的

    # 从下往上删除标记的行
    for row in sorted(rows_to_delete, reverse=True):
        sheet_preprocessing.delete_rows(row)

    # 收集处理后的A列数据
    processed_a_values = [sheet_preprocessing[f'A{row}'].value for row in range(3, sheet_preprocessing.max_row + 1) if sheet_preprocessing[f'A{row}'].value is not None]

    # 处理 TOPS_Costs url.xlsx
    wb_url = openpyxl.load_workbook(url_file)
    sheet_url = wb_url.active

    # 清除 TOPS_Costs url.xlsx 中 A 列的数据
    for row in range(3, sheet_url.max_row + 1):
        sheet_url[f'A{row}'].value = None

    # 将处理后的数据写入 TOPS_Costs url.xlsx
    for index, value in enumerate(processed_a_values, start=3):
        sheet_url[f'A{index}'].value = value

    # 保存修改后的 TOPS_Costs url.xlsx
    wb_url.save(url_file)
    print(f"处理完成。TOPS_Costs url.xlsx 已更新。")

    # 重新打开 TOPS_Costs url.xlsx，这次使用 data_only=True 来获取计算后的值
    wb_url_values = openpyxl.load_workbook(url_file, data_only=True)
    sheet_url_values = wb_url_values.active

    # 收集 B 列数据（对应 A 列不为空的行）
    b_column_data = []
    for row in range(3, sheet_url_values.max_row + 1):
        if sheet_url_values[f'A{row}'].value is not None:
            b_value = sheet_url_values[f'B{row}'].value
            if b_value is not None:
                b_column_data.append(str(b_value))

    # 更新 readurl.txt 文件
    with open(readurl_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(b_column_data))
    print(f"处理完成。readurl.txt 已更新。")

# 使用示例
if __name__ == "__main__":
    preprocessing_file = r"C:\data\TOPS_Costs preprocessing.xlsx"
    url_file = r"C:\data\TOPS_Costs url.xlsx"
    readurl_file = r"C:\data\readurl.txt"
    process_tops_costs_file(preprocessing_file, url_file, readurl_file)